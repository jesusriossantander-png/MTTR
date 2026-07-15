# -*- coding: utf-8 -*-
"""Extrae los valores YA CALCULADOS de la hoja MTTR del Excel exportado de Google Sheets.

El dashboard muestra literalmente los números de la hoja MTTR (Reparaci.,
Reparaci. Represent, MTTR por prioridad 0/1/2 y MTTR total, en días y en horas
donde existan), mes por mes. No se reimplementa ninguna fórmula: si la hoja
cambia su criterio, el dashboard la sigue.

También cuenta los equipos "en taller" (filas de IMPRESION sin fecha de egreso).

Uso:  python tools/parse_mttr.py planilla.xlsx [data.js]
"""
import json, re, sys, unicodedata
import openpyxl
from openpyxl.utils import range_boundaries

MES_NUM = {'ENERO':1,'FEBRERO':2,'MARZO':3,'ABRIL':4,'MAYO':5,'JUNIO':6,'JULIO':7,
           'AGOSTO':8,'SEPTIEMBRE':9,'SETIEMBRE':9,'OCTUBRE':10,'NOVIEMBRE':11,'DICIEMBRE':12}
ORDEN_TIPOS = ['B-OH2','B-OH4','B-BB1','B-BB2','B-BB3','B-BB4','B-BB5','B-VS1','B-VS5','B-VS6',
               'B-TORN','B-PROG','B-DOSIF','B-NEU','B-TRIP','B-VAP','B-SUND','B-AUX','B-ROT',
               'T','C','R','V','VTF','VARIOS','E','I']

def sin_acentos(s):
    return unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode()

def parse_mes(texto):
    """'enero 2026' / 'febrero 25' / 'ENERO--- 2024' / 'SETIEMBRE 2025' -> 'YYYY-MM'."""
    t = sin_acentos(str(texto)).upper()
    m = re.match(r'^\s*([A-Z]+)\s*[- ]*\s*(\d{2,4})\s*$', t)
    if not m or m.group(1) not in MES_NUM:
        return None
    y = int(m.group(2))
    if y < 100: y += 2000
    return f'{y}-{MES_NUM[m.group(1)]:02d}'

def num(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return round(float(v), 1)
    try: return round(float(str(v).replace(',', '.')), 1)
    except ValueError: return 0

def fila_tipo(v):
    t = str(v or '').strip().upper()
    return t if t in ORDEN_TIPOS else None

def main():
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else 'data.js'
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb['MTTR']

    # 1) Bloques definidos como TABLAS (2025 y 2026): el mes está como FECHA en el
    #    encabezado (o en la fila de arriba), y la unidad en el texto de las
    #    columnas MTTR: '(días)' o '(HORAS)'.
    from datetime import datetime, date as fecha
    def mes_de_fila(r, c1, c2):
        for rr in (r, r - 1):
            for c in range(c1, c2 + 1):
                v = ws.cell(row=rr, column=c).value
                if isinstance(v, (datetime, fecha)):
                    return f'{v.year}-{v.month:02d}'
                m = parse_mes(v) if isinstance(v, str) else None
                if m:
                    return m
        return None

    bloques = []   # (fila, col, mes, metrica, filas)
    for ref in ws.tables.values():
        ref = ref if isinstance(ref, str) else ref.ref
        c1, r1, c2, r2 = range_boundaries(ref)
        mes = mes_de_fila(r1, c1, c2)
        if not mes:
            continue
        hdr = ' '.join(str(ws.cell(row=r1, column=c).value or '') for c in range(c1, c2 + 1)).upper()
        if 'REPARACI' not in hdr:
            continue
        metrica = 'h' if 'HORAS' in hdr else 'd'
        filas = []
        for r in range(r1 + 1, r2 + 1):
            t = fila_tipo(ws.cell(row=r, column=c1).value)
            if t:
                filas.append((t, [num(ws.cell(row=r, column=c1 + i).value) for i in range(1, 7)]))
        if filas:
            bloques.append((r1, c1, mes, metrica, filas))

    # 2) Bloques de 2024 (sin tabla): encabezado 'ENERO--- 2024' en la fila,
    #    tipos en col A, valores en B..G (izquierda) y J..O (derecha).
    for row in ws.iter_rows(min_row=1, max_row=700):
        for cell in row:
            mes = parse_mes(cell.value) if isinstance(cell.value, str) else None
            if not (mes and mes.startswith('2024')):
                continue
            base = 1 if cell.column <= 8 else 9   # bloque izq: valores B..G; der: J..O
            filas = []
            r = cell.row + 2
            while r <= ws.max_row:
                t = fila_tipo(ws.cell(row=r, column=1).value)
                if not t:
                    break
                filas.append((t, [num(ws.cell(row=r, column=base + i).value) for i in range(1, 7)]))
                r += 1
            if filas:
                bloques.append((cell.row, base, mes, 'd', filas))

    # 3) Barrido genérico: bloques sueltos no definidos como tabla (p. ej. febrero
    #    2025). Celda con mes + 'Reparaci.' a la derecha + tipos debajo en la
    #    misma columna.
    vistos = {(b[2], b[3]) for b in bloques}
    filas_tabla = set()
    for b in bloques:
        for i in range(28):
            filas_tabla.add((b[0] + i, b[1]))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            v = cell.value
            mes = None
            if isinstance(v, (datetime, fecha)):
                mes = f'{v.year}-{v.month:02d}'
            elif isinstance(v, str):
                mes = parse_mes(v)
            if not mes or (cell.row, cell.column) in filas_tabla:
                continue
            hr = None
            for cand in (cell.row, cell.row + 1):
                htxt = ' '.join(str(ws.cell(row=cand, column=cell.column + i).value or '') for i in range(1, 7)).upper()
                if 'REPARACI' in htxt:
                    hr, hdr = cand, htxt
                    break
            if hr is None:
                continue
            metrica = 'h' if 'HORAS' in hdr else 'd'
            if (mes, metrica) in vistos:
                continue
            filas = []
            r = hr + 1
            while r <= ws.max_row:
                t = fila_tipo(ws.cell(row=r, column=cell.column).value)
                if not t:
                    break
                filas.append((t, [num(ws.cell(row=r, column=cell.column + i).value) for i in range(1, 7)]))
                r += 1
            if len(filas) >= 5:
                bloques.append((cell.row, cell.column, mes, metrica, filas))
                vistos.add((mes, metrica))

    meses = {}
    for r1, c1, mes, metrica, filas in sorted(bloques, key=lambda b: (b[2], b[0], b[1])):
        destino = meses.setdefault(mes, {})
        if metrica in destino:
            continue
        destino[metrica] = {t: vals for t, vals in filas}

    # Meses sin ningún dato (todo cero) no aportan: se descartan.
    def tiene_datos(md):
        return any(any(v) for v in md.get('d', {}).values())
    meses = {m: md for m, md in meses.items() if tiene_datos(md)}

    # 3) En taller: filas de IMPRESION con fecha de ingreso y sin egreso real.
    #    La planilla espejo solo trae la hoja MTTR (agregados); si no hay
    #    IMPRESION, el conteo "en taller" queda en 0 y el resto sigue igual.
    wip, wip_tipos = 0, {}
    if 'IMPRESION' in wb.sheetnames:
        wsi = wb['IMPRESION']
        for row in wsi.iter_rows(min_row=2, values_only=True):
            if len(row) < 12:
                continue
            fing, real = row[3], row[6]
            tag = str(row[0] or '').strip()
            if not tag or fing is None or str(fing).strip() in ('', '-'):
                continue
            if real is None or str(real).strip() in ('', '-'):
                tipo = str(row[11] or 'S/D').strip().upper() or 'S/D'
                wip += 1
                wip_tipos[tipo] = wip_tipos.get(tipo, 0) + 1

    if len(meses) < 12:
        print(f'ERROR: solo {len(meses)} meses detectados; se aborta para no perder datos.')
        sys.exit(1)
    data = {'order': ORDEN_TIPOS, 'months': meses, 'wip': wip,
            'wipTipos': dict(sorted(wip_tipos.items(), key=lambda kv: -kv[1]))}
    with open(out, 'w', encoding='utf-8') as f:
        f.write('const MTTR_DATA=' + json.dumps(data, ensure_ascii=False, separators=(',', ':')) + ';')
    horas = sum(1 for md in meses.values() if 'h' in md)
    print(f'OK: {len(meses)} meses ({horas} con horas), {wip} en taller -> {out}')

if __name__ == '__main__':
    main()
